from __future__ import annotations

import csv
import io
import logging
import random
import time
from dataclasses import asdict
from typing import BinaryIO, Callable, Iterable, TypeVar

from models import ReviewRecord

T = TypeVar("T")


def setup_logging(level: str = "INFO") -> logging.Logger:
    logger = logging.getLogger("ymaps_reviews")
    logger.setLevel(getattr(logging, level.upper(), logging.INFO))

    if not logger.handlers:
        handler = logging.StreamHandler()
        formatter = logging.Formatter(
            fmt="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        handler.setFormatter(formatter)
        logger.addHandler(handler)

    logger.propagate = False
    return logger


def sleep_random(min_seconds: float, max_seconds: float) -> None:
    time.sleep(random.uniform(min_seconds, max_seconds))


def normalize_whitespace(value: str | None) -> str:
    if value is None:
        return ""
    return " ".join(value.replace("\xa0", " ").split()).strip()


def unique_non_empty(items: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []

    for item in items:
        value = item.strip()
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)

    return result


def retry_call(
    func: Callable[[], T],
    *,
    attempts: int,
    delay_seconds: float,
    backoff: float,
    logger: logging.Logger | None = None,
    retry_exceptions: tuple[type[BaseException], ...] = (Exception,),
) -> T:
    current_delay = delay_seconds
    last_exc: BaseException | None = None

    for attempt in range(1, attempts + 1):
        try:
            return func()
        except retry_exceptions as exc:
            last_exc = exc
            if attempt >= attempts:
                break

            if logger is not None:
                logger.warning(
                    "Попытка %s/%s завершилась ошибкой: %s. Повтор через %.1f сек.",
                    attempt,
                    attempts,
                    exc,
                    current_delay,
                )
            time.sleep(current_delay)
            current_delay *= backoff

    assert last_exc is not None
    raise last_exc


def reviews_to_csv_bytes(records: list[ReviewRecord]) -> bytes:
    buffer = io.StringIO()
    fieldnames = [
        "residential_complex_input",
        "ymaps_card_name",
        "ymaps_card_address",
        "ymaps_card_url",
        "review_date",
        "user_name",
        "review_text",
    ]

    writer = csv.DictWriter(buffer, fieldnames=fieldnames, delimiter=";")
    writer.writeheader()

    for record in records:
        writer.writerow(asdict(record))

    return buffer.getvalue().encode("utf-8-sig")


def decode_uploaded_text_file(uploaded_file: BinaryIO | None) -> str:
    if uploaded_file is None:
        return ""

    raw = uploaded_file.read()
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue

    return raw.decode("utf-8", errors="ignore")

from __future__ import annotations

import csv
import io
import logging
import random
import time
from dataclasses import asdict
from typing import BinaryIO, Callable, Iterable, TypeVar

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from models import ReviewRecord

T = TypeVar("T")


def setup_logging(level: str = "INFO") -> logging.Logger:
    logger = logging.getLogger("ymaps_reviews")
    logger.setLevel(getattr(logging, level.upper(), logging.INFO))

    if not logger.handlers:
        handler = logging.StreamHandler()
        formatter = logging.Formatter(
            fmt="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        handler.setFormatter(formatter)
        logger.addHandler(handler)

    logger.propagate = False
    return logger


def sleep_random(min_seconds: float, max_seconds: float) -> None:
    time.sleep(random.uniform(min_seconds, max_seconds))


def normalize_whitespace(value: str | None) -> str:
    if value is None:
        return ""
    return " ".join(value.replace("\xa0", " ").split()).strip()


def unique_non_empty(items: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []

    for item in items:
        value = item.strip()
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)

    return result


def retry_call(
    func: Callable[[], T],
    *,
    attempts: int,
    delay_seconds: float,
    backoff: float,
    logger: logging.Logger | None = None,
    retry_exceptions: tuple[type[BaseException], ...] = (Exception,),
) -> T:
    current_delay = delay_seconds
    last_exc: BaseException | None = None

    for attempt in range(1, attempts + 1):
        try:
            return func()
        except retry_exceptions as exc:
            last_exc = exc
            if attempt >= attempts:
                break

            if logger is not None:
                logger.warning(
                    "Попытка %s/%s завершилась ошибкой: %s. Повтор через %.1f сек.",
                    attempt,
                    attempts,
                    exc,
                    current_delay,
                )
            time.sleep(current_delay)
            current_delay *= backoff

    assert last_exc is not None
    raise last_exc


def reviews_to_csv_bytes(records: list[ReviewRecord]) -> bytes:
    buffer = io.StringIO()
    fieldnames = [
        "residential_complex_input",
        "ymaps_card_name",
        "ymaps_card_address",
        "ymaps_card_url",
        "review_date",
        "user_name",
        "review_text",
    ]

    writer = csv.DictWriter(buffer, fieldnames=fieldnames, delimiter=";")
    writer.writeheader()

    for record in records:
        writer.writerow(asdict(record))

    return buffer.getvalue().encode("utf-8-sig")


def reviews_to_xlsx_bytes(records: list[ReviewRecord]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reviews"

    headers = [
        "residential_complex_input",
        "ymaps_card_name",
        "ymaps_card_address",
        "ymaps_card_url",
        "review_date",
        "user_name",
        "review_text",
    ]

    worksheet.append(headers)

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for record in records:
        row = asdict(record)
        worksheet.append([
            row["residential_complex_input"],
            row["ymaps_card_name"],
            row["ymaps_card_address"],
            row["ymaps_card_url"],
            row["review_date"],
            row["user_name"],
            row["review_text"],
        ])

    column_widths = {
        "A": 35,
        "B": 35,
        "C": 45,
        "D": 60,
        "E": 18,
        "F": 25,
        "G": 100,
    }

    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def decode_uploaded_text_file(uploaded_file: BinaryIO | None) -> str:
    if uploaded_file is None:
        return ""

    raw = uploaded_file.read()
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue

    return raw.decode("utf-8", errors="ignore")
